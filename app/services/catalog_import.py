from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime

import csv
import io

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import models


# ----------------------------
# Helpers
# ----------------------------

def _norm_header(h: str) -> str:
    """Normalize header strings to snake-ish keys."""
    if h is None:
        return ""
    h = str(h).strip().lower()
    for ch in ["-", "/", "\\", ".", "(", ")", "[", "]", "{", "}", ":"]:
        h = h.replace(ch, " ")
    h = h.replace("%", " pct ")
    h = " ".join(h.split())
    h = h.replace(" ", "_")
    return h


def _to_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # remove currency symbols/commas
    s = s.replace("$", "").replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(v: Any) -> Optional[int]:
    f = _to_float(v)
    if f is None:
        return None
    return int(round(f))


def _to_bool(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ["true", "yes", "y", "1", "active"]:
        return True
    if s in ["false", "no", "n", "0", "inactive"]:
        return False
    return None


def _pick(d: Dict[str, Any], *keys: str) -> Any:
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return None


@dataclass
class ImportSummary:
    rows_seen: int = 0
    rows_skipped: int = 0
    vendors_created: int = 0
    items_created: int = 0
    items_updated: int = 0
    prices_logged: int = 0
    errors: List[str] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


# ----------------------------
# Row Mapping (flexible)
# ----------------------------

def _map_row(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    Accepts many column name variants and returns a canonical row dict.

    Required:
      - item_name (or name/product)
      - unit_price (or price/cost)

    Optional:
      - vendor_name
      - sku
      - category
      - unit
      - description
      - lead_time_days
      - min_order_qty
      - is_active
    """
    item_name = _pick(raw, "item_name", "name", "product", "product_name", "title")
    vendor_name = _pick(raw, "vendor", "vendor_name", "supplier", "supplier_name", "store", "source_vendor")
    unit_price = _pick(raw, "unit_price", "price", "cost", "unit_cost", "real_unit_cost", "your_cost", "actual_cost")
    unit = _pick(raw, "unit", "uom", "unit_of_measure", "measure")
    sku = _pick(raw, "sku", "item_sku", "part_number", "part_no")
    category = _pick(raw, "category", "cat", "type")
    description = _pick(raw, "description", "desc", "details", "notes")
    lead_time_days = _pick(raw, "lead_time_days", "lead_time", "lead_days", "supplier_lead_time_days")
    min_order_qty = _pick(raw, "min_order_qty", "moq", "minimum_order", "minimum_quantity")
    is_active = _pick(raw, "is_active", "active", "status")

    return {
        "vendor_name": (str(vendor_name).strip() if vendor_name is not None else None),
        "name": (str(item_name).strip() if item_name is not None else None),
        "unit_price": _to_float(unit_price),
        "unit": (str(unit).strip() if unit is not None else None),
        "sku": (str(sku).strip() if sku is not None else None),
        "category": (str(category).strip() if category is not None else None),
        "description": (str(description).strip() if description is not None else None),
        "lead_time_days": _to_int(lead_time_days),
        "min_order_qty": _to_float(min_order_qty),
        "is_active": _to_bool(is_active),
    }


# ----------------------------
# DB Upsert Logic
# ----------------------------

def _get_or_create_vendor(db: Session, vendor_name: str, summary: ImportSummary) -> models.Vendor:
    vendor_name = (vendor_name or "Unknown Vendor").strip()

    v = db.query(models.Vendor).filter(models.Vendor.name == vendor_name).first()
    if v:
        return v

    v = models.Vendor(name=vendor_name)
    db.add(v)
    db.flush()  # get v.id without commit
    summary.vendors_created += 1
    return v


def _find_item(db: Session, vendor_id: int, name: str, sku: Optional[str]) -> Optional[models.CatalogItem]:
    q = db.query(models.CatalogItem).filter(models.CatalogItem.vendor_id == vendor_id)

    if sku:
        found = q.filter(models.CatalogItem.sku == sku).first()
        if found:
            return found

    # fallback: match by vendor + name
    return q.filter(models.CatalogItem.name == name).first()


def _upsert_item(db: Session, vendor: models.Vendor, mapped: Dict[str, Any], summary: ImportSummary,
                 default_category: Optional[str], source: str, note: Optional[str],
                 log_history_on_same_price: bool) -> None:

    name = mapped["name"]
    if not name:
        summary.rows_skipped += 1
        summary.errors.append("Row skipped: missing item name")
        return

    unit_price = mapped["unit_price"]
    if unit_price is None:
        summary.rows_skipped += 1
        summary.errors.append(f"Row skipped: missing/invalid unit_price for item '{name}'")
        return

    sku = mapped.get("sku")
    item = _find_item(db, vendor.id, name, sku)

    now = datetime.utcnow()
    category = mapped.get("category") or default_category

    if item is None:
        item = models.CatalogItem(
            vendor_id=vendor.id,
            name=name,
            description=mapped.get("description"),
            sku=sku,
            category=category,
            unit=mapped.get("unit") or "each",
            unit_price=float(unit_price),
            lead_time_days=mapped.get("lead_time_days"),
            min_order_qty=mapped.get("min_order_qty"),
            is_active=(mapped.get("is_active") if mapped.get("is_active") is not None else True),
            last_updated_at=now,
        )
        db.add(item)
        db.flush()
        summary.items_created += 1

        # always log first price
        h = models.CatalogPriceHistory(
            catalog_item_id=item.id,
            price=float(unit_price),
            source=source,
            note=note or "Imported initial price",
            recorded_at=now,
        )
        db.add(h)
        summary.prices_logged += 1
        return

    # update fields (merge)
    changed = False

    # price change check
    old_price = float(item.unit_price or 0.0)
    new_price = float(unit_price)

    price_changed = abs(old_price - new_price) > 0.00001
    if price_changed:
        item.unit_price = new_price
        item.last_updated_at = now
        changed = True

        h = models.CatalogPriceHistory(
            catalog_item_id=item.id,
            price=new_price,
            source=source,
            note=note or "Imported price update",
            recorded_at=now,
        )
        db.add(h)
        summary.prices_logged += 1
    else:
        if log_history_on_same_price:
            h = models.CatalogPriceHistory(
                catalog_item_id=item.id,
                price=new_price,
                source=source,
                note=note or "Imported (same price)",
                recorded_at=now,
            )
            db.add(h)
            summary.prices_logged += 1

    # merge non-price fields only if provided (don’t overwrite with blanks)
    if mapped.get("description"):
        item.description = mapped["description"]
        changed = True

    if mapped.get("unit"):
        item.unit = mapped["unit"]
        changed = True

    if sku and (item.sku != sku):
        item.sku = sku
        changed = True

    if category and (item.category != category):
        item.category = category
        changed = True

    if mapped.get("lead_time_days") is not None:
        item.lead_time_days = mapped["lead_time_days"]
        changed = True

    if mapped.get("min_order_qty") is not None:
        item.min_order_qty = mapped["min_order_qty"]
        changed = True

    if mapped.get("is_active") is not None:
        item.is_active = bool(mapped["is_active"])
        changed = True

    if changed:
        summary.items_updated += 1


# ----------------------------
# Public Import Functions
# ----------------------------

def import_csv_bytes(
    db: Session,
    content: bytes,
    source: str = "csv_import",
    note: Optional[str] = None,
    default_vendor: Optional[str] = None,
    default_category: Optional[str] = None,
    log_history_on_same_price: bool = False,
) -> ImportSummary:

    summary = ImportSummary()

    text = content.decode("utf-8", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        summary.errors.append("CSV has no headers")
        return summary

    # normalize headers
    norm_fields = [_norm_header(h) for h in reader.fieldnames]

    for row in reader:
        summary.rows_seen += 1
        raw = {}
        for original, norm in zip(reader.fieldnames, norm_fields):
            raw[norm] = row.get(original)

        mapped = _map_row(raw)
        vendor_name = mapped.get("vendor_name") or default_vendor or "Unknown Vendor"
        vendor = _get_or_create_vendor(db, vendor_name, summary)

        try:
            _upsert_item(
                db=db,
                vendor=vendor,
                mapped=mapped,
                summary=summary,
                default_category=default_category,
                source=source,
                note=note,
                log_history_on_same_price=log_history_on_same_price,
            )
        except Exception as e:
            summary.errors.append(f"Row error (CSV) item='{mapped.get('name')}': {e}")

    db.commit()
    return summary


def import_xlsx_bytes(
    db: Session,
    content: bytes,
    sheet_name: Optional[str] = None,
    source: str = "excel_import",
    note: Optional[str] = None,
    default_vendor: Optional[str] = None,
    default_category: Optional[str] = None,
    log_history_on_same_price: bool = False,
) -> ImportSummary:

    summary = ImportSummary()

    wb = load_workbook(io.BytesIO(content), data_only=True)

    ws = None
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            summary.errors.append(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return summary
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]  # first sheet

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        summary.errors.append("Excel sheet is empty")
        return summary

    headers = rows[0]
    norm_headers = [_norm_header(h) for h in headers]

    if not any(norm_headers):
        summary.errors.append("Excel headers look empty/invalid")
        return summary

    for r in rows[1:]:
        # Skip totally empty row
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue

        summary.rows_seen += 1
        raw: Dict[str, Any] = {}
        for idx, key in enumerate(norm_headers):
            if not key:
                continue
            raw[key] = r[idx] if idx < len(r) else None

        mapped = _map_row(raw)
        vendor_name = mapped.get("vendor_name") or default_vendor or "Unknown Vendor"
        vendor = _get_or_create_vendor(db, vendor_name, summary)

        try:
            _upsert_item(
                db=db,
                vendor=vendor,
                mapped=mapped,
                summary=summary,
                default_category=default_category,
                source=source,
                note=note,
                log_history_on_same_price=log_history_on_same_price,
            )
        except Exception as e:
            summary.errors.append(f"Row error (XLSX) item='{mapped.get('name')}': {e}")

    db.commit()
    return summary

def apply_price_updates_from_xlsx_bytes(
    db: Session,
    content: bytes,
    sheet_name: str = "PRICE_UPDATES",
    source: str = "excel_price_updates",
    note: Optional[str] = None,
    default_vendor: Optional[str] = None,
    log_history_on_same_price: bool = False,
) -> ImportSummary:
    """
    Reads PRICE_UPDATES sheet:
      vendor_name | item_name* | sku | new_unit_price* | source | note | effective_date

    - Finds matching catalog item by (vendor + sku) or (vendor + item_name)
    - Updates unit_price if changed
    - Always logs to price history when changed (and optionally when same)
    """
    summary = ImportSummary()

    wb = load_workbook(io.BytesIO(content), data_only=True)

    if sheet_name not in wb.sheetnames:
        summary.errors.append(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return summary

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        summary.errors.append("PRICE_UPDATES sheet is empty")
        return summary

    headers = rows[0]
    norm_headers = [_norm_header(h) for h in headers]

    # Normalize expected columns
    # vendor_name, item_name, sku, new_unit_price, source, note, effective_date
    for r in rows[1:]:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue

        summary.rows_seen += 1

        raw: Dict[str, Any] = {}
        for idx, key in enumerate(norm_headers):
            if not key:
                continue
            raw[key] = r[idx] if idx < len(r) else None

        vendor_name = _pick(raw, "vendor_name", "vendor", "supplier_name") or default_vendor or "Unknown Vendor"
        item_name = _pick(raw, "item_name", "name", "product", "product_name", "title")
        sku = _pick(raw, "sku", "item_sku", "part_number", "part_no")
        new_price = _pick(raw, "new_unit_price", "unit_price", "price", "cost", "unit_cost")

        mapped_price = _to_float(new_price)
        if not item_name and not sku:
            summary.rows_skipped += 1
            summary.errors.append("Row skipped: missing item_name and sku")
            continue
        if mapped_price is None:
            summary.rows_skipped += 1
            summary.errors.append(f"Row skipped: invalid new_unit_price for '{item_name or sku}'")
            continue

        vendor = _get_or_create_vendor(db, str(vendor_name), summary)

        # Find item
        item = _find_item(db, vendor.id, str(item_name).strip() if item_name else "", str(sku).strip() if sku else None)
        if not item:
            summary.rows_skipped += 1
            summary.errors.append(f"Row skipped: item not found for vendor='{vendor.name}', name='{item_name}', sku='{sku}'")
            continue

        now = datetime.utcnow()
        old_price = float(item.unit_price or 0.0)
        new_price_val = float(mapped_price)

        row_source = _pick(raw, "source") or source
        row_note = _pick(raw, "note") or note or "Excel PRICE_UPDATES upload"

        price_changed = abs(old_price - new_price_val) > 0.00001
        if price_changed:
            item.unit_price = new_price_val
            item.last_updated_at = now
            summary.items_updated += 1

            h = models.CatalogPriceHistory(
                catalog_item_id=item.id,
                price=new_price_val,
                source=str(row_source),
                note=str(row_note),
                recorded_at=now,
            )
            db.add(h)
            summary.prices_logged += 1
        else:
            if log_history_on_same_price:
                h = models.CatalogPriceHistory(
                    catalog_item_id=item.id,
                    price=new_price_val,
                    source=str(row_source),
                    note=str(row_note) + " (same price)",
                    recorded_at=now,
                )
                db.add(h)
                summary.prices_logged += 1

    db.commit()
    return summary
